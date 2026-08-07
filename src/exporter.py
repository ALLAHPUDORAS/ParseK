import csv
import json
import logging
import os
from typing import List, Dict, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import (
    OUTPUT_DIR,
    DEFAULT_EXPORT_CSV,
    DEFAULT_EXPORT_JSON,
    DEFAULT_EXPORT_RAW_CSV,
    DEFAULT_EXPORT_RAW_JSON,
    DEFAULT_EXPORT_RAW_XLSX,
    DEFAULT_EXPORT_TEXT,
)

logger = logging.getLogger("LeadPipeline.Exporter")

class Exporter:
    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = os.path.abspath(output_dir)
        os.makedirs(self.output_dir, exist_ok=True)
        logger.info("Exporter initialized. Output directory: %s", self.output_dir)

    def purge_previous_exports(self) -> None:
        for filename in [
            DEFAULT_EXPORT_JSON,
            DEFAULT_EXPORT_CSV,
            DEFAULT_EXPORT_RAW_JSON,
            DEFAULT_EXPORT_RAW_CSV,
            DEFAULT_EXPORT_RAW_XLSX,
            "leads_formatted.xlsx",
            "leads_report.html",
        ]:
            path = os.path.join(self.output_dir, filename)
            try:
                if os.path.exists(path):
                    os.remove(path)
                    logger.info("Removed stale output file: %s", path)
            except Exception as exc:
                logger.exception("Failed to remove stale output file %s: %s", path, exc)

    def _get_contact_data(self, lead: Dict[str, Any]) -> Dict[str, Any]:
        contacts = lead.get("contacts") or lead.get("raw_contacts") or {}
        if not isinstance(contacts, dict):
            return {}
        return contacts

    def _flatten_contacts(self, lead: Dict[str, Any]) -> Dict[str, Any]:
        contacts = self._get_contact_data(lead)
        flattened: Dict[str, Any] = {}
        for contact_type in ["telegram", "emails", "discord", "linkedin", "twitter_x", "skype", "other_socials"]:
            values = contacts.get(contact_type, [])
            if isinstance(values, str):
                values = [values]
            flattened[contact_type] = ", ".join(str(value).strip() for value in values if str(value).strip())
        return flattened

    def _write_excel_report(self, leads: List[Dict[str, Any]], filename: str = "leads_formatted.xlsx") -> str:
        path = os.path.join(self.output_dir, filename)
        try:
            rows = []
            for lead in leads:
                flattened = self._flatten_contacts(lead)
                row = {
                    "name": lead.get("name", ""),
                    "vertical": lead.get("vertical", ""),
                    "geo": lead.get("geo", ""),
                    "website": lead.get("website", ""),
                    "status": lead.get("status", ""),
                    "source": lead.get("source", ""),
                    "telegram_status": "active" if flattened.get("telegram") else "none",
                }
                row.update(flattened)
                rows.append(row)

            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame(columns=["name", "vertical", "geo", "website", "status", "source", "telegram_status", "telegram", "emails", "discord", "linkedin", "twitter_x", "skype", "other_socials"])

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Leads", index=False)
                workbook = writer.book
                sheet = writer.sheets["Leads"]

                header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
                header_font = Font(bold=True)
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for column_cells in sheet.columns:
                    max_length = 12
                    for cell in column_cells:
                        if cell.value is None:
                            continue
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    adjusted_width = min(max(12, max_length + 2), 60)
                    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

                sheet.freeze_panes = "A2"

            logger.info("Exported %d leads to Excel: %s", len(leads), path)
        except Exception as exc:
            logger.exception("Failed to generate Excel report %s: %s", path, exc)
        return path

    def _write_raw_excel_report(self, leads: List[Dict[str, Any]], filename: str = DEFAULT_EXPORT_RAW_XLSX) -> str:
        path = os.path.join(self.output_dir, filename)
        try:
            rows = []
            for lead in leads:
                raw_contacts = self._get_contact_data(lead)
                row = {
                    "name": lead.get("name", ""),
                    "vertical": lead.get("vertical", ""),
                    "geo": lead.get("geo", ""),
                    "website": lead.get("website", ""),
                    "status": lead.get("status", ""),
                    "source": lead.get("source", ""),
                    "raw_telegram": ", ".join(raw_contacts.get("telegram", [])) if isinstance(raw_contacts.get("telegram", []), list) else str(raw_contacts.get("telegram", "")),
                    "raw_emails": ", ".join(raw_contacts.get("emails", [])) if isinstance(raw_contacts.get("emails", []), list) else str(raw_contacts.get("emails", "")),
                    "raw_discord": ", ".join(raw_contacts.get("discord", [])) if isinstance(raw_contacts.get("discord", []), list) else str(raw_contacts.get("discord", "")),
                    "raw_linkedin": ", ".join(raw_contacts.get("linkedin", [])) if isinstance(raw_contacts.get("linkedin", []), list) else str(raw_contacts.get("linkedin", "")),
                    "raw_twitter_x": ", ".join(raw_contacts.get("twitter_x", [])) if isinstance(raw_contacts.get("twitter_x", []), list) else str(raw_contacts.get("twitter_x", "")),
                    "raw_skype": ", ".join(raw_contacts.get("skype", [])) if isinstance(raw_contacts.get("skype", []), list) else str(raw_contacts.get("skype", "")),
                    "raw_other_socials": ", ".join(raw_contacts.get("other_socials", [])) if isinstance(raw_contacts.get("other_socials", []), list) else str(raw_contacts.get("other_socials", "")),
                }
                rows.append(row)

            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame(columns=[
                    "name", "vertical", "geo", "website", "status", "source",
                    "raw_telegram", "raw_emails", "raw_discord", "raw_linkedin",
                    "raw_twitter_x", "raw_skype", "raw_other_socials",
                ])

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Raw Leads", index=False)
                workbook = writer.book
                sheet = writer.sheets["Raw Leads"]

                header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
                header_font = Font(bold=True)
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for column_cells in sheet.columns:
                    max_length = 12
                    for cell in column_cells:
                        if cell.value is None:
                            continue
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    adjusted_width = min(max(12, max_length + 2), 60)
                    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

                sheet.freeze_panes = "A2"

            logger.info("Exported %d raw leads to Excel: %s", len(leads), path)
        except Exception as exc:
            logger.exception("Failed to generate raw Excel report %s: %s", path, exc)
        return path

    def _write_html_report(self, leads: List[Dict[str, Any]], filename: str = "leads_report.html") -> str:
        path = os.path.join(self.output_dir, filename)
        try:
            rows = []
            for lead in leads:
                row = {
                    "name": lead.get("name", ""),
                    "vertical": lead.get("vertical", ""),
                    "geo": lead.get("geo", ""),
                    "website": lead.get("website", ""),
                    "status": lead.get("status", ""),
                    "source": lead.get("source", ""),
                }
                row.update(self._flatten_contacts(lead))
                rows.append(row)

            escaped_rows = []
            for row in rows:
                escaped_rows.append({key: (value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") if isinstance(value, str) else value) for key, value in row.items()})

            columns = ["name", "vertical", "geo", "website", "status", "source", "telegram", "emails", "discord", "linkedin", "twitter_x", "skype", "other_socials"]
            table_rows = []
            for row in escaped_rows:
                cells = "".join(f"<td>{row.get(column, '')}</td>" for column in columns)
                table_rows.append(f"<tr>{cells}</tr>")

            html = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\" />
  <title>Leads Report</title>
  <link rel=\"stylesheet\" href=\"https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css\" />
  <style>body{{font-family:Arial,sans-serif;padding:16px;}} h1{{margin-bottom:8px;}} .controls{{margin-bottom:12px;}} table{{width:100%;}} .dataTables_filter{{float:right;}}</style>
</head>
<body>
  <h1>Leads Report</h1>
  <div class=\"controls\"><input id=\"lead-search\" type=\"search\" placeholder=\"Search leads...\" /></div>
  <table id=\"leads-table\" class=\"display\">
    <thead><tr>{''.join(f'<th>{column}</th>' for column in columns)}</tr></thead>
    <tbody>{''.join(table_rows)}</tbody>
  </table>
  <script src=\"https://code.jquery.com/jquery-3.7.1.min.js\"></script>
  <script src=\"https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js\"></script>
  <script>
    $(document).ready(function() {{
      $('#leads-table').DataTable({{searching:true, paging:true, ordering:true, pageLength:25}});
      $('#lead-search').on('keyup', function() {{
        $('#leads-table').DataTable().search(this.value).draw();
      }});
    }});
  </script>
</body>
</html>"""
            with open(path, "w", encoding="utf-8") as handle:
                handle.write(html)
            logger.info("Exported %d leads to HTML report: %s", len(leads), path)
        except Exception as exc:
            logger.exception("Failed to generate HTML report %s: %s", path, exc)
        return path

    def _lead_key(self, lead: Dict[str, Any]) -> tuple:
        name = (lead.get("name") or "").strip().lower()
        website = (lead.get("website") or "").strip().lower()
        return name, website

    def _dedupe_leads(self, leads: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        unique: List[Dict[str, Any]] = []
        seen = set()
        last_key = None
        for lead in leads:
            key = self._lead_key(lead)
            if key in seen or key == last_key:
                continue
            seen.add(key)
            unique.append(lead)
            last_key = key
        return unique

    def save_leads(self, leads: List[Dict[str, Any]], filename_json: str = DEFAULT_EXPORT_JSON, filename_csv: str = DEFAULT_EXPORT_CSV) -> Dict[str, str]:
        deduped = self._dedupe_leads(leads)
        self.to_json(deduped, filename=filename_json)
        self.to_csv(deduped, filename=filename_csv)
        self._write_excel_report(deduped)
        self._write_html_report(deduped)
        return {
            "json": os.path.join(self.output_dir, filename_json),
            "csv": os.path.join(self.output_dir, filename_csv),
            "excel": os.path.join(self.output_dir, "leads_formatted.xlsx"),
            "html": os.path.join(self.output_dir, "leads_report.html"),
        }

    def save_raw_leads(self, leads: List[Dict[str, Any]], filename_json: str = DEFAULT_EXPORT_RAW_JSON, filename_csv: str = DEFAULT_EXPORT_RAW_CSV, filename_xlsx: str = DEFAULT_EXPORT_RAW_XLSX) -> Dict[str, str]:
        self.to_json(leads, filename=filename_json)
        self.to_csv(leads, filename=filename_csv)
        self._write_raw_excel_report(leads, filename=filename_xlsx)
        return {
            "json": os.path.join(self.output_dir, filename_json),
            "csv": os.path.join(self.output_dir, filename_csv),
            "excel": os.path.join(self.output_dir, filename_xlsx),
        }

    def save_empty_export(self) -> None:
        """Ensure export files exist even when there are no valid leads."""
        self.save_leads([])

    def to_json(self, leads: List[Dict[str, Any]], filename: str = DEFAULT_EXPORT_JSON) -> str:
        path = os.path.join(self.output_dir, filename)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(leads, handle, ensure_ascii=False, indent=2)
        logger.info("Exported %d leads to JSON: %s", len(leads), path)
        return path

    def to_csv(self, leads: List[Dict[str, Any]], filename: str = DEFAULT_EXPORT_CSV) -> str:
        path = os.path.join(self.output_dir, filename)
        with open(path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["name", "vertical", "geo", "website", "status", "source", "contacts"])
            for lead in leads:
                contacts = []
                contact_data = self._get_contact_data(lead)
                for contact_type in ["emails", "telegram", "discord", "linkedin", "twitter_x", "skype", "other_socials"]:
                    values = contact_data.get(contact_type, [])
                    if isinstance(values, str):
                        values = [values]
                    if values:
                        contacts.append(f"{contact_type}:{','.join(str(value) for value in values)}")
                writer.writerow([
                    lead.get("name", ""),
                    lead.get("vertical", ""),
                    lead.get("geo", ""),
                    lead.get("website", ""),
                    lead.get("status", ""),
                    lead.get("source", ""),
                    " | ".join(contacts)
                ])
        logger.info("Exported %d leads to CSV: %s", len(leads), path)
        return path

    def to_text(self, leads: List[Dict[str, Any]], filename: str = DEFAULT_EXPORT_TEXT) -> str:
        path = os.path.join(self.output_dir, filename)
        lines = []
        for lead in leads:
            contact_data = self._get_contact_data(lead)
            contacts = []
            for contact_type in ["telegram", "skype", "emails", "discord", "linkedin", "twitter_x", "other_socials"]:
                values = contact_data.get(contact_type, [])
                if isinstance(values, str):
                    values = [values]
                if values:
                    contacts.extend(str(value) for value in values)
            contact_str = ", ".join(contacts)
            line = f"{lead.get('name', '')} | {lead.get('vertical', '')}/{lead.get('geo', '')} | Contact: {contact_str}"
            lines.append(line)

        with open(path, "w", encoding="utf-8") as handle:
            handle.write("\n".join(lines))

        logger.info("Exported %d leads to text: %s", len(leads), path)
        return path
